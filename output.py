import sys, json

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Color


def load_data():
    return json.load(sys.stdin)


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def export_excel():
    # data = load_data()

    wb = Workbook()
    ws = wb.active
    cell = ws['A1']

    cell.value = 12

    thin = Side(border_style="thin", color="000000")

    dark_red = PatternFill("solid", fgColor=Color(theme=5, tint=0.4))
    light_red = PatternFill("solid", fgColor=Color(theme=5, tint=0.6))
    bold = Font(b=True, color="000000")
    regular = Font(b=False, color="000000")


    styles = {
        'fill': PatternFill("solid", fgColor=Color(theme=5, tint=0.4)),
        'border': Border(top=thin, left=thin, right=thin, bottom=thin),
        'font': Font(b=True, color="000000")
    }

    style_range(ws, 'A1:C3', **styles)

    wb.save(sys.argv[1] + '.xlsx')

def export_json():
    data = load_data()
    with open(sys.argv[1] + '.json', 'w+') as f:
        json.dump(data, f, indent=4)

    # print March monthly JFA HTR DATC values 
    print(data['year']['quarters'][0]['months'][2]['data']['1070054'])


if __name__ == '__main__':
    export_excel()
